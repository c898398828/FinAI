import logging
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.database import get_db
from app.models.user import User
from app.schemas.financial import (
    FinancialCashflow,
    FinancialRecordCreate,
    FinancialRecordList,
    FinancialRecordOut,
    FinancialSummary,
    FinancialTrend,
    UploadResult,
)
from app.services.financial_service import (
    bulk_create_records,
    create_record,
    get_cashflow,
    get_records,
    get_summary,
    get_trend,
)
from app.utils.excel_parser import parse_excel

logger = logging.getLogger(__name__)
router = APIRouter()


def _ensure_company(current_user: User) -> int:
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="请先创建或加入企业")
    return current_user.company_id


@router.post("/upload", response_model=UploadResult)
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)

    if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="仅支持 Excel(.xlsx/.xls) 或 CSV 文件")

    content = await file.read()
    records, errors = parse_excel(content, file.filename)
    if not records and errors:
        return UploadResult(success=False, records_imported=0, errors=errors)

    count = bulk_create_records(db, company_id, current_user.id, records)
    return UploadResult(success=True, records_imported=count, errors=errors)


@router.get("/records", response_model=FinancialRecordList)
def list_records(
    category: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)
    items, total = get_records(db, company_id, category, start_date, end_date, skip, limit)
    return FinancialRecordList(total=total, items=items)


@router.post("/records", response_model=FinancialRecordOut)
def add_record(
    data: FinancialRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)
    return create_record(db, company_id, current_user.id, data)


@router.get("/summary", response_model=FinancialSummary)
def financial_summary(
    start_date: date | None = None,
    end_date: date | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)
    return get_summary(db, company_id, start_date, end_date)


@router.get("/trend", response_model=FinancialTrend)
def financial_trend(
    start_date: date | None = None,
    end_date: date | None = None,
    period: str = "month",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)
    if period not in {"month", "quarter"}:
        raise HTTPException(status_code=400, detail="period must be 'month' or 'quarter'")
    return get_trend(db, company_id, start_date, end_date, period)


@router.get("/cashflow", response_model=FinancialCashflow)
def financial_cashflow(
    start_date: date | None = None,
    end_date: date | None = None,
    period: str = "month",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    company_id = _ensure_company(current_user)
    if period not in {"month", "quarter"}:
        raise HTTPException(status_code=400, detail="period must be 'month' or 'quarter'")
    return get_cashflow(db, company_id, start_date, end_date, period)


@router.get("/template")
def download_template():
    """下载财务数据导入模板 (.xlsx)。"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "财务数据"

        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")

        required_font = Font(bold=True, size=12, color="FFFFFF")
        required_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        optional_font = Font(bold=True, size=12, color="FFFFFF")
        optional_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = [
            ("日期", True),
            ("类别", True),
            ("金额", True),
            ("子类别", False),
            ("描述", False),
        ]
        widths = [15, 12, 15, 18, 30]

        for idx, (name, required) in enumerate(headers, start=1):
            c = ws.cell(row=1, column=idx, value=name)
            c.font = required_font if required else optional_font
            c.fill = required_fill if required else optional_fill
            c.alignment = center
            c.border = thin
            ws.column_dimensions[chr(64 + idx)].width = widths[idx - 1]
        ws.row_dimensions[1].height = 28

        demo_rows = [
            ("2026-02-03", "收入", 50000.00, "销售收入", "2月产品销售"),
            ("2026-02-05", "支出", 12000.00, "办公费用", "办公场地租金"),
            ("2026-02-10", "支出", 36000.00, "员工工资", "2月工资发放"),
            ("2026-02-18", "资产", 120000.00, "现金", "期末账户余额"),
            ("2026-02-20", "负债", 30000.00, "应付账款", "供应商结算"),
        ]
        for r_idx, row in enumerate(demo_rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=c_idx, value=val)
                c.alignment = center
                c.border = thin

        notes = [
            "说明：",
            "1. 必填列：日期、类别、金额",
            "2. 日期格式：YYYY-MM-DD（例如 2026-02-15）",
            "3. 类别可选值：收入、支出、资产、负债",
            "4. 金额为正数，保留两位小数",
            "5. 子类别、描述为可选",
        ]
        note_font = Font(size=10, color="666666")
        for i, note in enumerate(notes):
            row = 9 + i
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=note)
            c.font = note_font
            c.alignment = Alignment(wrap_text=True)

        dv = DataValidation(
            type="list",
            formula1='"收入,支出,资产,负债"',
            allow_blank=False,
            errorTitle="类别错误",
            error="类别必须为：收入/支出/资产/负债",
        )
        dv.add("B2:B1000")
        ws.add_data_validation(dv)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="financial_import_template.xlsx"'},
        )
    except Exception as exc:
        logger.exception("模板生成失败: %s", exc)
        raise HTTPException(status_code=500, detail="模板生成失败")
