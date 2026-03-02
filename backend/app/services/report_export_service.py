from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.report import Report


HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
LABEL_FONT = Font(bold=True, color="334155")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CURRENCY_FORMAT = '¥#,##0.00'


def _safe_sheet_name(name: str) -> str:
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid:
        name = name.replace(ch, "_")
    name = name.strip() or "Sheet"
    return name[:31]


def _is_ai_section(key: str) -> bool:
    return "ai" in key.lower()


def _base_sheet(wb: Workbook, report: Report) -> int:
    ws = wb.active
    ws.title = "报表概览"

    ws["A1"] = "财务报表导出"
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = HEADER_FONT
    ws.merge_cells("A1:D1")

    meta = [
        ("报表标题", report.title or "-"),
        ("报表类型", report.report_type),
        ("开始日期", str(report.period_start)),
        ("结束日期", str(report.period_end)),
        ("生成时间", str(report.created_at)),
    ]

    row = 3
    for key, value in meta:
        ws.cell(row=row, column=1, value=key).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24

    return row


def _style_range(ws, max_row: int, min_col: int = 1, max_col: int = 2):
    for cells in ws.iter_rows(min_row=1, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in cells:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _write_overview_metrics(wb: Workbook, data_json: dict, start_row: int):
    ws = wb["报表概览"]

    row = start_row + 1
    ws.cell(row=row, column=1, value="核心指标").font = LABEL_FONT
    row += 1

    ws.cell(row=row, column=1, value="指标")
    ws.cell(row=row, column=2, value="金额")
    ws.cell(row=row, column=1).font = LABEL_FONT
    ws.cell(row=row, column=2).font = LABEL_FONT
    ws.cell(row=row, column=1).fill = SUB_HEADER_FILL
    ws.cell(row=row, column=2).fill = SUB_HEADER_FILL
    row += 1

    for key, value in data_json.items():
        if _is_ai_section(str(key)):
            continue
        if isinstance(value, (int, float)):
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=2, value=float(value))
            ws.cell(row=row, column=2).number_format = CURRENCY_FORMAT
            row += 1

    _style_range(ws, max(row, 8))


def _write_dict_sheet(wb: Workbook, sheet_name: str, payload: dict):
    sheet = wb.create_sheet(_safe_sheet_name(sheet_name))
    sheet["A1"] = sheet_name
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = HEADER_FONT
    sheet.merge_cells("A1:C1")

    sheet["A2"] = "项目"
    sheet["B2"] = "金额"
    sheet["A2"].font = LABEL_FONT
    sheet["B2"].font = LABEL_FONT
    sheet["A2"].fill = SUB_HEADER_FILL
    sheet["B2"].fill = SUB_HEADER_FILL

    row = 3
    total = 0.0
    for item_key, item_value in payload.items():
        amount = float(item_value or 0)
        sheet.cell(row=row, column=1, value=str(item_key))
        sheet.cell(row=row, column=2, value=amount)
        sheet.cell(row=row, column=2).number_format = CURRENCY_FORMAT
        total += amount
        row += 1

    if len(payload) > 1:
        sheet.cell(row=row, column=1, value="合计").font = LABEL_FONT
        sheet.cell(row=row, column=2, value=total).font = LABEL_FONT
        sheet.cell(row=row, column=2).number_format = CURRENCY_FORMAT

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 18
    _style_range(sheet, max(row, 3))


def build_report_excel(report: Report) -> BytesIO:
    wb = Workbook()
    data_json = report.data_json or {}

    row = _base_sheet(wb, report)
    _write_overview_metrics(wb, data_json, row)

    for key, value in data_json.items():
        if _is_ai_section(str(key)):
            continue
        if isinstance(value, dict):
            _write_dict_sheet(wb, str(key), value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
