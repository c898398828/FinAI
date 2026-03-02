import asyncio
from io import BytesIO
import unittest

import httpx
from openpyxl import load_workbook

from app.main import app


class FinancialTemplateApiTests(unittest.TestCase):
    async def _get_template(self) -> httpx.Response:
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
            return await client.get("/api/financial/template")

    def test_template_download_returns_excel(self) -> None:
        resp = asyncio.run(self._get_template())

        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp.headers.get("content-type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "attachment; filename=\"financial_import_template.xlsx\"",
            resp.headers.get("content-disposition", ""),
        )
        self.assertGreater(len(resp.content), 0)

    def test_template_excel_payload_is_readable(self) -> None:
        resp = asyncio.run(self._get_template())
        wb = load_workbook(filename=BytesIO(resp.content))

        self.assertIn("类别说明", wb.sheetnames)
        sheet = wb.active
        self.assertEqual(sheet.max_column, 5)
        self.assertTrue(sheet.cell(row=1, column=1).value)
        self.assertTrue(sheet.cell(row=1, column=2).value)
        self.assertTrue(sheet.cell(row=1, column=3).value)


if __name__ == "__main__":
    unittest.main()
